#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
耘宇数据源 TXT 转 Excel 脚本
功能：将包含物料、订单、库存变动记录的 JSON 文本文件转换为多 Sheet 的 Excel 文件
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def extract_json_blocks(content):
    """从文本中提取所有顶层 JSON 对象"""
    depth = 0
    blocks = []
    start = -1
    for i, ch in enumerate(content):
        if ch == '{':
            if depth == 0:
                start = i
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0 and start >= 0:
                blocks.append(content[start:i + 1])
                start = -1
    return blocks


def safe_get(d, key, default=''):
    """安全获取字典值"""
    if d is None:
        return default
    return d.get(key, default)


def flatten_material(row):
    """展平物料信息行"""
    unit_list = row.get('unit', []) or []
    unit_val = ', '.join([u.get('val', '') for u in unit_list if u])

    group = row.get('group') or {}
    crop_category = row.get('cropCategory') or {}
    unit_conv = row.get('unitConversionGroup') or {}
    sale_unit = row.get('saleUint') or {}

    custom_att = row.get('customAttMap') or {}
    custom_att_str = '; '.join([f'{k}: {v}' for k, v in custom_att.items()])

    return {
        '物料ID': row.get('id', ''),
        '基地ID': row.get('baseId', ''),
        '物料编号': row.get('materialNum', ''),
        '物料名称': row.get('materialName', ''),
        '作物分类': crop_category.get('name', '') if crop_category else '',
        '单位': unit_val,
        '物料分组': group.get('name', '') if group else '',
        '销售单位': sale_unit.get('val', '') if sale_unit else '',
        '销售单价': row.get('saleUnitPrice', ''),
        '单位换算组': unit_conv.get('name', '') if unit_conv else '',
        '自定义属性': custom_att_str,
        '状态': row.get('status', ''),
        '可删除': row.get('canDel', ''),
        '仓库ID': row.get('warehouseId', ''),
    }


def flatten_order(row):
    """展平订单主表信息行"""
    customer = row.get('customerInfo') or {}
    purpose = row.get('purposeVo') or {}

    status_map = {
        -1: '已取消',
        0: '待发货',
        1: '部分发货',
        2: '已发货',
        10: '待收款',
        11: '部分收款',
        12: '已收款',
        20: '已完成',
    }
    take_status_map = {0: '未发货', 1: '部分发货', 2: '已发货'}
    collect_status_map = {0: '未收款', 1: '部分收款', 2: '已收款'}

    return {
        '订单ID': row.get('orderId', ''),
        '订单编号': row.get('orderNumber', ''),
        '客户名称': customer.get('name', '') if customer else '',
        '客户编号': customer.get('number', '') if customer else '',
        '用途': purpose.get('purpose', '') if purpose else '',
        '计划发货日期': row.get('planDeliveryDate', ''),
        '最晚发货日期': row.get('latestDeliveryDate', ''),
        '发货状态': take_status_map.get(row.get('takeStatus', 0), str(row.get('takeStatus', ''))),
        '收款状态': collect_status_map.get(row.get('collectStatus', 0), str(row.get('collectStatus', ''))),
        '订单状态': status_map.get(row.get('status', 0), str(row.get('status', ''))),
        '取消原因': row.get('cancelReason', ''),
        '备注': row.get('remark', ''),
        '计划发货总量': row.get('totalPlanDelivery', ''),
        '已发货总量': row.get('totalDelivered', ''),
        '已收货总量': row.get('totalTake', ''),
        '订单金额': row.get('orderPrice', ''),
        '超发订单金额': row.get('overOrderPrice', ''),
        '已收款金额': row.get('collectAmount', ''),
        '明细数量': len(row.get('details', []) or []),
    }


def flatten_order_detail(order_row, detail):
    """展平订单明细行"""
    material = detail.get('materialVo') or {}
    unit_vo = detail.get('unitVo') or {}
    convert_plan = detail.get('convertPlanDelivery') or {}
    convert_over = detail.get('convertOverDelivery') or {}
    convert_take = detail.get('convertTakeNumber') or {}

    return {
        '订单编号': order_row.get('orderNumber', ''),
        '客户名称': (order_row.get('customerInfo') or {}).get('name', ''),
        '明细ID': detail.get('id', ''),
        '物料编号': material.get('materialNum', '') if material else '',
        '物料名称': material.get('materialName', '') if material else '',
        '计划发货量': detail.get('planDelivery', ''),
        '实际发货量': detail.get('overDelivery', ''),
        '单位': unit_vo.get('val', '') if unit_vo else '',
        '单价': detail.get('unitPrice', ''),
        '订单金额': detail.get('orderPrice', ''),
        '超发金额': detail.get('overOrderPrice', ''),
        '待发货量': detail.get('waitDelivery', ''),
        '已收货数量': detail.get('takeNumber', ''),
        '收货损耗': detail.get('takeDamage', ''),
        '损耗备注': detail.get('takeDamageRemark', ''),
        '明细备注': detail.get('detailRemark', ''),
        '换算计划发货(基本单位)': f"{convert_plan.get('right', '')} {convert_plan.get('middle', '')}" if convert_plan else '',
        '换算实际发货(基本单位)': f"{convert_over.get('right', '')} {convert_over.get('middle', '')}" if convert_over else '',
        '换算已收货(基本单位)': f"{convert_take.get('right', '')} {convert_take.get('middle', '')}" if convert_take else '',
    }


def flatten_inventory(row):
    """展平库存变动记录行"""
    change_type_map = {
        1: '入库',
        2: '出库',
        3: '调拨',
        4: '盘点',
        5: '调整',
    }

    return {
        '记录ID': row.get('id', ''),
        '记录编号': row.get('recordNo', ''),
        '库存ID': row.get('inventoryId', ''),
        '操作时间': row.get('operationTime', '').replace('T', ' ')[:19] if row.get('operationTime') else '',
        '变动类型': change_type_map.get(row.get('changeType', 0), str(row.get('changeType', ''))),
        '车牌号': row.get('vehicleNo', ''),
        '仓库ID': row.get('warehouseId', ''),
        '仓库名称': row.get('warehouseName', ''),
        '仓库编码': row.get('code', ''),
        '库位ID': row.get('warehouseLocationId', ''),
        '库位名称': row.get('warehouseLocationName', ''),
        '批次号': row.get('invBatch', ''),
        '基地ID': row.get('baseId', ''),
        '物料ID': row.get('materialId', ''),
        '物料编号': row.get('materialNum', ''),
        '物料名称': row.get('materialName', ''),
        '订单明细ID': row.get('orderDetailId', ''),
        '订单发货日期': row.get('orderDeliveryDate', ''),
        '变动数量': row.get('amount', ''),
        '基本单位数量': row.get('basicAmount', ''),
        '单位ID': row.get('unitId', ''),
        '单位': row.get('unit', ''),
        '数量描述': row.get('amountDesc', ''),
        '第二单位数量': row.get('secondaryAmount', ''),
        '第二单位': row.get('secondaryUnit', ''),
        '销售单位ID': row.get('saleUnitId', ''),
        '销售单位': row.get('saleUnit', ''),
        '变动日期': row.get('changeDate', '').replace('T', ' ')[:19] if row.get('changeDate') else '',
        '创建时间': row.get('createTime', '').replace('T', ' ')[:19] if row.get('createTime') else '',
        '更新时间': row.get('updateTime', '').replace('T', ' ')[:19] if row.get('updateTime') else '',
        '备注': row.get('remark', ''),
        '状态': row.get('status', ''),
        '已删除': row.get('deleted', ''),
    }


def write_sheet(ws, data_list, title):
    """将数据写入工作表并设置格式"""
    if not data_list:
        ws.cell(row=1, column=1, value='暂无数据')
        return

    # 获取所有列名
    headers = list(data_list[0].keys())

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for row_idx, row_data in enumerate(data_list, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border

    # 自动调整列宽
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row_data in data_list:
            value = str(row_data.get(header, ''))
            # 中文字符算2个宽度
            length = sum(2 if ord(c) > 127 else 1 for c in value)
            if length > max_length:
                max_length = length
        # 限制最大宽度
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置行高
    ws.row_dimensions[1].height = 25


def main():
    # 输入文件路径
    input_file = '耘宇数据源.txt'
    output_file = '耘宇数据源.xlsx'

    print(f'正在读取文件: {input_file}')

    # 读取文件
    with open(input_file, 'r', encoding='utf-8') as f:
        content = f.read()

    # 提取所有 JSON 块
    blocks = extract_json_blocks(content)
    print(f'共找到 {len(blocks)} 个 JSON 数据块')

    # 分类数据
    material_rows = []
    order_rows = []
    inventory_rows = []

    for i, block_str in enumerate(blocks):
        try:
            data = json.loads(block_str)
            rows = data.get('rows', [])
            total = data.get('total', 0)

            if not rows:
                continue

            # 根据第一条记录的字段判断数据类型
            first_row = rows[0]
            if 'recordNo' in first_row and 'changeType' in first_row:
                # 库存变动记录（优先判断，因为库存记录也有 materialNum/materialName）
                inventory_rows.extend(rows)
                print(f'  块{i}: 库存变动记录 - total={total}, 本块{len(rows)}条')
            elif 'orderId' in first_row and 'orderNumber' in first_row:
                # 订单信息
                order_rows.extend(rows)
                print(f'  块{i}: 订单信息 - total={total}, 本块{len(rows)}条')
            elif 'materialNum' in first_row and 'materialName' in first_row and 'group' in first_row:
                # 物料信息（通过 group 字段区分）
                material_rows.extend(rows)
                print(f'  块{i}: 物料信息 - total={total}, 本块{len(rows)}条')
            else:
                print(f'  块{i}: 未知数据类型, keys={list(first_row.keys())[:8]}')

        except json.JSONDecodeError as e:
            print(f'  块{i}: 解析失败 - {e}')

    print(f'\n数据汇总:')
    print(f'  物料信息: {len(material_rows)} 条')
    print(f'  订单信息: {len(order_rows)} 条')
    print(f'  库存变动记录: {len(inventory_rows)} 条')

    # 展平数据
    print('\n正在处理数据...')

    material_data = [flatten_material(row) for row in material_rows]
    order_data = [flatten_order(row) for row in order_rows]

    # 订单明细需要展开
    order_detail_data = []
    for order_row in order_rows:
        details = order_row.get('details', []) or []
        for detail in details:
            order_detail_data.append(flatten_order_detail(order_row, detail))

    inventory_data = [flatten_inventory(row) for row in inventory_rows]

    print(f'  物料信息处理完成: {len(material_data)} 条')
    print(f'  订单主表处理完成: {len(order_data)} 条')
    print(f'  订单明细处理完成: {len(order_detail_data)} 条')
    print(f'  库存变动记录处理完成: {len(inventory_data)} 条')

    # 创建 Excel 文件
    print(f'\n正在生成 Excel 文件: {output_file}')
    wb = Workbook()

    # 删除默认的 Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 创建各个 Sheet
    ws_material = wb.create_sheet('物料信息')
    write_sheet(ws_material, material_data, '物料信息')

    ws_order = wb.create_sheet('订单主表')
    write_sheet(ws_order, order_data, '订单主表')

    ws_order_detail = wb.create_sheet('订单明细')
    write_sheet(ws_order_detail, order_detail_data, '订单明细')

    ws_inventory = wb.create_sheet('库存变动记录')
    write_sheet(ws_inventory, inventory_data, '库存变动记录')

    # 保存文件
    wb.save(output_file)
    file_size = os.path.getsize(output_file)
    print(f'Excel 文件生成成功! 文件大小: {file_size / 1024:.1f} KB')
    print(f'文件路径: {os.path.abspath(output_file)}')

    # 输出各 Sheet 列名
    print('\n各 Sheet 字段说明:')
    print(f'  【物料信息】({len(material_data)}条): {", ".join(material_data[0].keys()) if material_data else "无数据"}')
    print(f'  【订单主表】({len(order_data)}条): {", ".join(order_data[0].keys()) if order_data else "无数据"}')
    print(f'  【订单明细】({len(order_detail_data)}条): {", ".join(order_detail_data[0].keys()) if order_detail_data else "无数据"}')
    print(f'  【库存变动记录】({len(inventory_data)}条): {", ".join(inventory_data[0].keys()) if inventory_data else "无数据"}')


if __name__ == '__main__':
    main()
